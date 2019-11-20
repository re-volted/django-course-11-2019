# pip install openpyxl
import json

from django.core import serializers
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.shortcuts import render, render_to_response

# Create your views here.
from django.urls import reverse
from django.utils import timezone
from django.views import generic
from django.views.decorators.http import require_http_methods

from books.forms import CommentForm
from .models import Book


def books_list(req):
    books = Book.objects.all()
    page = req.GET.get('page')
    paginator = Paginator(books, 20)
    books = paginator.get_page(page)

    format = req.GET.get('format')
    if format == 'json':
        books = serializers.serialize('json', books)
        books = json.loads(books)
        return  JsonResponse(books, safe=False)

    elif format == 'xml':
        books = serializers.serialize('xml', books)
        return  HttpResponse(books, content_type='application/xhtml+xml')

    elif format == 'yaml':
        books = serializers.serialize('yaml', books)
        return  HttpResponse(books, content_type='application/xhtml+xml')

    return render(
        request=req,
        template_name="books/list.html",
        context={
            'books': books,
        }
    )

# to use it, needs to change "books" to "page_obj" in paginator class div of list.html
class ListView(generic.ListView):
    model = Book
    context_object_name = "books"
    template_name = "books/list.html"
    paginate_by = 20

    def get_queryset(self):
        return self.model.objects.all()


class DetailView(generic.DetailView, generic.FormView):
    model = Book
    template_name = "books/details.html"
    form_class = CommentForm

# POST books/1/comment
@require_http_methods(["POST"])
def book_comment(req, pk):
    form = CommentForm(data=req.POST)
    if form.is_valid():
        comment = form.save(commit=False)
        book = Book.objects.get(pk=pk)
        comment.book = book
        comment.save()
        messages.add_message(req, messages.INFO, "A comment was added")
        return HttpResponseRedirect(reverse('books:detail', args=(pk,)))
    else:
        messages.add_message(req, messages.ERROR, "An error occurred while adding comment.")
        return HttpResponseRedirect(reverse('books:detail', args=(pk,)))

def books_to_excel(req):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    response['Content-Disposition'] = "attachment; filename=List_{date}.xlsx".format(date=timezone.now().strftime("%Y-%m-%d"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Books list"

    books = Book.objects.all()
    page = req.GET.get('page')
    paginator = Paginator(books, 20)
    books = paginator.get_page(page)

    columns = [
        "Book title", "Authors", "Book price", "Book rating", "Book publisher name", "Book publication date"
    ]

    ws.append(columns)


    for book in books:
        row = [
            book.title,
            (", ").join([f"{author.first_name} {author.last_name}" for author in book.authors.all()]),
            book.price,
            book.rating,
            book.publisher.name,
            book.pub_date
        ]
        ws.append(row)

    ft = Font(bold=True)
    first_row_cells = [ws["{l}1".format(l=letter)] for letter in 'ABCDEF']

    for cell in first_row_cells:
        cell.font = ft

    wb.save(response)
    return response